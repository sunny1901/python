from kazoo.client import KazooClient

from urllib import parse
from urllib.parse import quote, unquote

# 创建Zookeeper客户端
zk = KazooClient(hosts='dubboprd.sinosafe.com.cn:2181')

# 连接到Zookeeper服务器
zk.start()

# 读取节点数据
data, stat = zk.get("/")

print("节点数据：", data)
print("节点状态：", stat)

nodes = zk.get_children("/dubbo/com.sinosafe.demo.facade.UmDomainFacade/providers")
# print('>>',nodes)
for i in nodes:
    # print(i)
    data, stat = zk.get("/")
    # print("节点数据：", data)
    # print("节点状态：", stat)

# 关闭Zookeeper客户端连接

nodes = zk.get_children("/dubbo")
print('>> nodes ',len(nodes))

app = {}
app_none= {}
app['app_none'] = app_none

service_list = {}



pop = 0
for i in nodes:
    # print(i)
    _path = f"/dubbo/{i}"
    # data, stat = zk.get(_path)
    # print(data , stat)
    # # print("\t节点数据：", data, "节点状态：", stat )
    ch_nodes1 = zk.get_children(_path)
    # print( ch_nodes1 )
    app_none[i] = i
    if not service_list.get(i):
        service_list[i] = 1
    else:
        service_list[i] = service_list[i] + 1

    for j in ch_nodes1:
        # print('\t' , j)
        ch_nodes2 = zk.get_children(_path + "/" + j )
        for k in ch_nodes2:
    #         #print('\t\t' , k)
            url = unquote(k)
            result = parse.urlparse(url)
            ip = result.netloc
            queryParam = parse.parse_qs(result.query)
            if queryParam.get('application'):
                for app_name in queryParam.get('application'):

                    app_info = app.get(app_name)
                    if not app_info:
                        app_info = {}
                        app[app_name] = app_info

                    app_role = app_info.get(j) # providers , consumers
                    if not app_role:
                        app_role = {}
                        app_info[j] = app_role

                    app_role[i] = ip

                    interface = queryParam.get('interface')
                    for tmp_i in interface:
                        app_role[tmp_i] = ip


                    # app_dict['_none_app1'](i)
                    # assert i == 'webservice.provider.paymentquery.facade.FINReceiveCMISInfoService'
                    if app_none.get(i):
                        app_none.pop(i)
                        pop = pop + 1


## None ####################

print( 'App None 1 ' , len(app_none) )
for i in list(app_none):
    # print(i)
    _path = f"/dubbo/{i}"
    # data, stat = zk.get(_path)
    # print(data , stat)
    # # print("\t节点数据：", data, "节点状态：", stat )
    ch_nodes1 = zk.get_children(_path)
    # print( ch_nodes1 )

    for j in ch_nodes1:
        # print('\t' , j)
        ch_nodes2 = zk.get_children(_path + "/" + j )

        if j ==  'configurators':

            if len(ch_nodes2) == 0 :
                print("Path None" , _path + "/" + j )
                if app_none.get(i):
                    app_none.pop(i)
                    pop = pop + 1
                continue

            app_info = app.get('None')
            if not app_info:
                app_info = {}
                app['None'] = app_info

            app_role = app_info.get(j)  # providers , consumers
            if not app_role:
                app_role = {}
                app_info[j] = app_role

            app_role[i] = ip
            interface = queryParam.get('interface')
            for tmp_i in interface:
                app_role[tmp_i] = ip




print( 'App None 2 ' , len(app_none) )


zk.stop()

print( '>>' , len(service_list) )
# for i in service_list:
    # print( '>>>>' ,i ,service_list[i] )

print(len(i))
count = 0

sync_list = ["cn.com.sinosafe.lina.shortdomain.dubbo.service.ShortDomainDubboService",
"cn.com.sinosafe.core.query.common.service.CoreClaimDubService",
"cn.com.sinosafe.clmautoevaluate.dubbo.DayuDubService",
"cn.com.sinosafe.cancelcase.api.dubbo.CaseDubboService",
"cn.com.sinosafe.claim.ZBCFservice.service.ZBCFSmsService",
"cn.com.sinosafe.report.base.service.GcRegistService",
"cn.com.sinosafe.claim.ZBCFservice.service.ZBCFCenterService",
"cn.com.sinosafe.claim.selfservice.admin.service.ReDispatchService",
"cn.com.sinosafe.fee.dubbo.service.SalesProductReceiveService",
"cn.com.sinosafe.fee.api.http.service.FeeForXBService",
"cn.com.sinosafe.fee.api.http.service.FeeForDSService",
"cn.com.sinosafe.fee.dubbo.service.CostManageProviderService",
"cn.com.sinosafe.tps.module.api.dubbo.provider.service.ITpsForClaimProviderService",
"cn.com.sinosafe.antifraud.dubbo.IAntiDubService"]


data = [
    ['App','Role' , 'DubboService' , '-']
]


for app_name in app:

    # app
    # print('app.', app_name)
    if app_name != 'None':
        continue

    # app.info
    app_info = app[app_name]
    for role_name in app_info:

        # print('\t', '', role_name )
        app_role = app_info[role_name]
        for _service_name in app_role:
            count = count + 1
            # print('\t\t' , _service_name )
            row = None
            if _service_name in sync_list:
                print( app_name , ',' , role_name  , ',', _service_name , ',' , '已迁移')
                row = [ app_name , role_name ,  _service_name , '已迁移' ]
            else:
                print(app_name, ',', role_name, ',', _service_name, ',' , '-')
                row = [app_name, role_name, _service_name, '-']
            data.append(row)
    else:
        count = count + 1
#

import csv
with open('data.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(data)

print('CSV 文件已生成')

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

# 将二维数组转换为 DataFrame
df = pd.DataFrame(data)


book = load_workbook('template.xlsx')
writer = pd.ExcelWriter('output2.xlsx', engine='openpyxl')
writer.book = book

# 将 DataFrame 写入到 Excel 文件中
df.to_excel(writer, sheet_name='Sheet1', index=False)

sheet = book['Sheet1']

worksheet = book["Sheet"]
book.remove(worksheet)
#
# # 遍历所有行，进行行合并
for i in range(1, len(df)):
    if df.iloc[i][0] == df.iloc[i - 1][0]:
        tmp = 2
        # mrege_i = i
        # while df.iloc[mrege_i][0] == df.iloc[i - 1][0]:
        #     tmp = tmp + 1
        #     mrege_i = mrege_i +1
        sheet.merge_cells(start_row=i + 1, end_row=i + tmp, start_column=1, end_column=1)
    else:
        tmp = -1
        sheet.cell(row=i + 1, column=1).alignment = Alignment(vertical='top')

print('CSV 文件已生成')

# 关闭 Excel 写入器
writer.save()
writer.close()
print('CSV 文件已生成')



print('>> App None ', len(app_none))

